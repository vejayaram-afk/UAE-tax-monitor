#!/usr/bin/env python3
"""
UAE Tax & Reporting Update Monitor
==================================

Checks the Federal Tax Authority, Ministry of Finance and IFRS Foundation for
new publications, tags them by topic, and builds a static dashboard.

Usage
-----
    python monitor.py              Run a check and rebuild the dashboard
    python monitor.py --check      Test every source URL and report status
    python monitor.py --rebuild    Rebuild the dashboard from stored data only
    python monitor.py --digest     Print the new-items digest (for email)
    python monitor.py --export     Write an Excel file of everything captured

Configuration lives in config.yaml. You do not need to edit this file.
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from urllib.parse import quote_plus, urljoin, urlparse

import requests
import yaml
from bs4 import BeautifulSoup

try:
    import feedparser
except ImportError:
    feedparser = None

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "config.yaml")
DATA_DIR = os.path.join(HERE, "data")
DOCS_DIR = os.path.join(HERE, "docs")
STORE_PATH = os.path.join(DATA_DIR, "items.json")

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/122.0 Safari/537.36")

# Link text this short, or matching these patterns, is navigation, not content.
NOISE = re.compile(
    r"^(home|about|contact|login|sign in|register|search|menu|more|read more|"
    r"next|previous|back|download|click here|view all|en|ar|english|arabic|"
    r"privacy|terms|sitemap|cookies|accessibility|faq|careers|share)$",
    re.I,
)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def load_store():
    if os.path.exists(STORE_PATH):
        with open(STORE_PATH, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return {}


def save_store(store):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(STORE_PATH, "w", encoding="utf-8") as fh:
        json.dump(store, fh, indent=2, ensure_ascii=False)


def item_id(url, title):
    """Stable identity for an item. URL first; title as a fallback."""
    basis = (url or "").split("#")[0].strip().lower() or (title or "").strip().lower()
    return hashlib.sha1(basis.encode("utf-8")).hexdigest()[:16]


def title_key(title, authority):
    """
    Near-duplicate key. The same announcement carried by six news outlets has
    near-identical headlines; this collapses them to one entry per authority
    tier, so an official notice and the press coverage of it both survive but
    six versions of the press coverage do not.
    """
    t = re.sub(r"[^a-z0-9 ]", "", (title or "").lower())
    t = re.sub(r"\b(the|a|an|of|for|to|in|on|and|uae|until|till)\b", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return hashlib.sha1(f"{authority}|{t[:90]}".encode("utf-8")).hexdigest()[:16]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def clean(text):
    return re.sub(r"\s+", " ", (text or "")).strip()


def categorise(title, categories):
    """Return every category whose keywords appear in the title."""
    low = (title or "").lower()
    hits = [name for name, words in categories.items()
            if any(w.lower() in low for w in words)]
    return hits or ["Other"]


def looks_like_content(text, href):
    if not href or href.startswith(("javascript:", "mailto:", "tel:", "#")):
        return False
    text = clean(text)
    if len(text) < 18:
        return False
    if NOISE.match(text):
        return False
    return True


def fetch(url, timeout, session):
    resp = session.get(url, timeout=timeout, headers={"User-Agent": UA})
    resp.raise_for_status()
    return resp


# ---------------------------------------------------------------------------
# collectors — each returns a list of dicts
# ---------------------------------------------------------------------------

def collect_html(src, cfg, session):
    url = src["url"]
    resp = fetch(url, cfg["settings"]["timeout"], session)
    soup = BeautifulSoup(resp.text, "html.parser")

    selector = src.get("selector") or "a"
    nodes = soup.select(selector)
    if not nodes:                       # selector missed — fall back to all links
        nodes = soup.find_all("a")

    out, seen = [], set()
    for node in nodes:
        href = node.get("href")
        text = clean(node.get_text())
        if not looks_like_content(text, href):
            continue
        full = urljoin(url, href)
        if full in seen:
            continue
        seen.add(full)
        out.append({"title": text, "url": full, "published": None})
    return out


def collect_rss(src, cfg, session):
    if feedparser is None:
        raise RuntimeError("feedparser is not installed — run: pip install feedparser")
    resp = fetch(src["url"], cfg["settings"]["timeout"], session)
    feed = feedparser.parse(resp.content)
    out = []
    for e in feed.entries:
        published = None
        if getattr(e, "published_parsed", None):
            published = datetime(*e.published_parsed[:6], tzinfo=timezone.utc).isoformat()
        out.append({
            "title": clean(getattr(e, "title", "")),
            "url": getattr(e, "link", ""),
            "published": published,
            "summary": clean(BeautifulSoup(getattr(e, "summary", ""), "html.parser").get_text())[:400],
        })
    return out


def collect_google_news(src, cfg, session):
    q = quote_plus(src["query"])
    url = f"https://news.google.com/rss/search?q={q}&hl=en-AE&gl=AE&ceid=AE:en"
    return collect_rss({"url": url}, cfg, session)


COLLECTORS = {"html": collect_html, "rss": collect_rss, "google_news": collect_google_news}


# ---------------------------------------------------------------------------
# main run
# ---------------------------------------------------------------------------

def run(cfg, store):
    session = requests.Session()
    settings = cfg["settings"]
    categories = cfg["categories"]
    new_items, report = [], []

    # index of near-duplicate titles already held, so repeated coverage of one
    # announcement does not fill the dashboard
    tkeys = {rec.get("title_key") for rec in store.values() if rec.get("title_key")}

    for src in cfg["sources"]:
        if not src.get("enabled", True):
            continue
        name = src["name"]
        try:
            items = COLLECTORS[src["type"]](src, cfg, session)
            added = 0
            for it in items:
                if not it.get("url"):
                    continue
                key = item_id(it["url"], it["title"])
                if key in store:
                    store[key]["last_seen"] = now_iso()
                    continue

                tkey = title_key(it["title"], src.get("authority", "news"))
                if tkey in tkeys:
                    continue
                tkeys.add(tkey)

                record = {
                    "id": key,
                    "title_key": tkey,
                    "title": it["title"],
                    "url": it["url"],
                    "summary": it.get("summary", ""),
                    "source": name,
                    "authority": src.get("authority", "news"),
                    "categories": categorise(it["title"], categories),
                    "published": it.get("published"),
                    "first_seen": now_iso(),
                    "last_seen": now_iso(),
                    "domain": urlparse(it["url"]).netloc.replace("www.", ""),
                    "read": False,
                    "actioned": False,
                }
                store[key] = record
                new_items.append(record)
                added += 1
            report.append((name, "OK", f"{len(items)} found, {added} new"))
        except Exception as exc:                                    # noqa: BLE001
            report.append((name, "FAIL", str(exc)[:110]))
        time.sleep(settings.get("delay", 1.5))

    # prune very old records
    cutoff = datetime.now(timezone.utc) - timedelta(days=settings.get("max_age_days", 400))
    for key in list(store):
        try:
            if datetime.fromisoformat(store[key]["first_seen"]) < cutoff:
                del store[key]
        except Exception:                                            # noqa: BLE001
            pass

    return new_items, report


def check_sources(cfg):
    session = requests.Session()
    print(f"\n  Checking {len(cfg['sources'])} sources\n  " + "-" * 62)
    for src in cfg["sources"]:
        name = src["name"]
        if not src.get("enabled", True):
            print(f"  SKIP  {name}")
            continue
        try:
            items = COLLECTORS[src["type"]](src, cfg, session)
            status = "OK  " if items else "EMPTY"
            print(f"  {status}  {name}  ({len(items)} items)")
            if not items:
                print(f"        -> check the URL or selector in config.yaml")
        except Exception as exc:                                     # noqa: BLE001
            print(f"  FAIL  {name}")
            print(f"        -> {str(exc)[:100]}")
        time.sleep(1)
    print("  " + "-" * 62 + "\n")


# ---------------------------------------------------------------------------
# dashboard
# ---------------------------------------------------------------------------

def build_dashboard(cfg, store):
    settings = cfg["settings"]
    new_days = settings.get("new_for_days", 7)
    threshold = datetime.now(timezone.utc) - timedelta(days=new_days)

    items = list(store.values())

    def sort_key(it):
        return it.get("published") or it.get("first_seen") or ""

    items.sort(key=sort_key, reverse=True)

    for it in items:
        try:
            it["is_new"] = datetime.fromisoformat(it["first_seen"]) > threshold
        except Exception:                                            # noqa: BLE001
            it["is_new"] = False

    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(os.path.join(DOCS_DIR, "data.json"), "w", encoding="utf-8") as fh:
        json.dump({"generated": now_iso(), "items": items}, fh, indent=2, ensure_ascii=False)

    cat_names = list(cfg["categories"].keys()) + ["Other"]
    html = HTML_TEMPLATE.replace("__TITLE__", settings.get("site_title", "Tax Monitor"))
    html = html.replace("__SUBTITLE__", settings.get("site_subtitle", ""))
    html = html.replace("__GENERATED__", datetime.now(timezone.utc).strftime("%d %B %Y, %H:%M UTC"))
    html = html.replace("__CATEGORIES__", json.dumps(cat_names))
    html = html.replace("__NEWDAYS__", str(new_days))
    html = html.replace("__ITEMS__", json.dumps(items, ensure_ascii=False))

    with open(os.path.join(DOCS_DIR, "index.html"), "w", encoding="utf-8") as fh:
        fh.write(html)
    return len(items)


# ---------------------------------------------------------------------------
# digest + email
# ---------------------------------------------------------------------------

def build_digest(new_items, cfg):
    if not new_items:
        return None
    by_cat = {}
    for it in new_items:
        for c in it["categories"]:
            by_cat.setdefault(c, []).append(it)

    order = list(cfg["categories"].keys()) + ["Other"]
    lines = [f"UAE TAX & REPORTING UPDATES — {datetime.now().strftime('%d %B %Y')}",
             "=" * 62, "", f"{len(new_items)} new item(s) found.", ""]
    for cat in order:
        if cat not in by_cat:
            continue
        lines.append(cat.upper())
        lines.append("-" * len(cat))
        for it in by_cat[cat][:20]:
            flag = "[OFFICIAL] " if it["authority"] == "official" else ""
            lines.append(f"  {flag}{it['title']}")
            lines.append(f"     {it['url']}")
            lines.append(f"     Source: {it['source']}")
            lines.append("")
        lines.append("")
    lines.append("-" * 62)
    lines.append("Automated monitor. Verify every position against the primary")
    lines.append("FTA / Ministry of Finance / IFRS text before advising a client.")
    return "\n".join(lines)


def send_email(digest, new_count):
    """Sends the digest if SMTP environment variables are set. Silent if not."""
    import smtplib
    from email.mime.text import MIMEText

    host = os.environ.get("SMTP_HOST")
    if not host or not digest:
        return False
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER")
    pwd = os.environ.get("SMTP_PASS")
    to = os.environ.get("EMAIL_TO", user)

    msg = MIMEText(digest, "plain", "utf-8")
    msg["Subject"] = f"UAE tax updates — {new_count} new item(s)"
    msg["From"] = user
    msg["To"] = to
    try:
        with smtplib.SMTP(host, port, timeout=30) as s:
            s.starttls()
            s.login(user, pwd)
            s.send_message(msg)
        print(f"  Email sent to {to}")
        return True
    except Exception as exc:                                         # noqa: BLE001
        print(f"  Email failed: {exc}")
        return False


def export_excel(store):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  openpyxl not installed — run: pip install openpyxl")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Updates"
    headers = ["First seen", "Published", "Category", "Authority", "Title",
               "Source", "Domain", "URL", "Reviewed?", "Client impact", "Action"]
    widths = [18, 18, 22, 12, 70, 34, 22, 60, 12, 34, 34]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5C8A")
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    items = sorted(store.values(), key=lambda x: x.get("first_seen", ""), reverse=True)
    for r, it in enumerate(items, start=2):
        row = [it.get("first_seen", "")[:10], (it.get("published") or "")[:10],
               ", ".join(it.get("categories", [])), it.get("authority", ""),
               it.get("title", ""), it.get("source", ""), it.get("domain", ""),
               it.get("url", ""), "", "", ""]
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=(i == 5), vertical="top")
    out = os.path.join(DOCS_DIR, "tax_updates_export.xlsx")
    os.makedirs(DOCS_DIR, exist_ok=True)
    wb.save(out)
    print(f"  Exported {len(items)} items to {out}")


# ---------------------------------------------------------------------------
# dashboard template
# ---------------------------------------------------------------------------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
  :root{
    --navy:#1f3864; --mid:#2e5c8a; --light:#dce6f1; --ink:#1a1a1a;
    --muted:#6b7280; --line:#e3e7ee; --bg:#f6f8fb; --new:#c0392b; --ok:#1e7e34;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;
       background:var(--bg);color:var(--ink);line-height:1.55;font-size:15px}
  header{background:var(--navy);color:#fff;padding:26px 28px}
  header h1{font-size:23px;font-weight:700;letter-spacing:-.2px}
  header p{font-size:13px;opacity:.82;margin-top:4px}
  .meta{font-size:12px;opacity:.7;margin-top:10px}
  .wrap{max-width:1180px;margin:0 auto;padding:22px 28px 70px}
  .stats{display:flex;gap:12px;flex-wrap:wrap;margin:20px 0}
  .stat{background:#fff;border:1px solid var(--line);border-radius:8px;
        padding:14px 18px;min-width:120px}
  .stat .n{font-size:25px;font-weight:700;color:var(--navy);line-height:1.1}
  .stat .l{font-size:11px;text-transform:uppercase;letter-spacing:.6px;
           color:var(--muted);margin-top:3px}
  .controls{background:#fff;border:1px solid var(--line);border-radius:8px;
            padding:16px 18px;margin-bottom:20px}
  .row{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:10px}
  .row:last-child{margin-bottom:0}
  .lab{font-size:11px;text-transform:uppercase;letter-spacing:.6px;
       color:var(--muted);min-width:74px;font-weight:600}
  button.f{background:#fff;border:1px solid var(--line);border-radius:20px;
           padding:5px 14px;font-size:13px;cursor:pointer;color:var(--ink);
           font-family:inherit;transition:.12s}
  button.f:hover{border-color:var(--mid)}
  button.f.on{background:var(--navy);color:#fff;border-color:var(--navy)}
  input[type=search]{flex:1;min-width:220px;padding:8px 12px;font-size:14px;
       border:1px solid var(--line);border-radius:6px;font-family:inherit}
  .item{background:#fff;border:1px solid var(--line);border-left:3px solid var(--line);
        border-radius:7px;padding:15px 18px;margin-bottom:9px;transition:.12s}
  .item:hover{box-shadow:0 2px 10px rgba(31,56,100,.08)}
  .item.official{border-left-color:var(--navy)}
  .item.news{border-left-color:#b8c4d6}
  .item.commentary{border-left-color:#8fa8c8}
  .item.isnew{background:#fffdf5}
  .item h3{font-size:15.5px;font-weight:600;line-height:1.4;margin-bottom:6px}
  .item h3 a{color:var(--navy);text-decoration:none}
  .item h3 a:hover{text-decoration:underline}
  .sum{font-size:13.5px;color:#444;margin-bottom:8px}
  .tags{display:flex;gap:6px;flex-wrap:wrap;align-items:center;font-size:11.5px}
  .tag{background:var(--light);color:var(--navy);padding:2px 9px;border-radius:11px;
       font-weight:600}
  .badge{background:var(--new);color:#fff;padding:2px 8px;border-radius:11px;
         font-weight:700;letter-spacing:.4px;font-size:10.5px}
  .badge.off{background:var(--ok)}
  .src{color:var(--muted)}
  .empty{background:#fff;border:1px solid var(--line);border-radius:8px;
         padding:44px;text-align:center;color:var(--muted)}
  footer{max-width:1180px;margin:0 auto;padding:0 28px 40px;font-size:12px;
         color:var(--muted);line-height:1.7}
  .warn{background:#fff8e6;border:1px solid #f0d9a0;border-radius:7px;
        padding:12px 16px;font-size:13px;margin-bottom:20px;color:#6b5320}
  @media print{.controls,.stats,button{display:none}body{background:#fff}}
</style>
</head>
<body>
<header>
  <h1>__TITLE__</h1>
  <p>__SUBTITLE__</p>
  <div class="meta">Last updated __GENERATED__</div>
</header>

<div class="wrap">
  <div class="warn">
    <strong>Automated monitor.</strong> Items are collected by machine and are not
    reviewed before publication. Verify every position against the primary FTA,
    Ministry of Finance or IFRS Foundation text before advising a client.
  </div>

  <div class="stats" id="stats"></div>

  <div class="controls">
    <div class="row">
      <span class="lab">Topic</span>
      <div id="catFilters"></div>
    </div>
    <div class="row">
      <span class="lab">Source</span>
      <button class="f on" data-auth="all">All</button>
      <button class="f" data-auth="official">Official only</button>
      <button class="f" data-auth="news">News</button>
      <button class="f" data-auth="commentary">Commentary</button>
      <button class="f" data-new="1">New only</button>
    </div>
    <div class="row">
      <span class="lab">Search</span>
      <input type="search" id="q" placeholder="Filter by keyword, e.g. e-invoicing, small business relief, IFRS 18">
    </div>
  </div>

  <div id="list"></div>
</div>

<footer>
  Generated automatically. Official sources are the Federal Tax Authority, the UAE
  Ministry of Finance and the IFRS Foundation; news and commentary items are third-party
  reports and may be inaccurate or incomplete. Nothing here is tax or accounting advice.
</footer>

<script>
const ITEMS = __ITEMS__;
const CATS = __CATEGORIES__;
const NEWDAYS = __NEWDAYS__;
let fCat = "all", fAuth = "all", fNew = false, fQ = "";

const catBox = document.getElementById('catFilters');
["all"].concat(CATS).forEach(c => {
  const b = document.createElement('button');
  b.className = 'f' + (c === 'all' ? ' on' : '');
  b.textContent = c === 'all' ? 'All' : c;
  b.dataset.cat = c;
  catBox.appendChild(b);
});

function esc(s){const d=document.createElement('div');d.textContent=s||'';return d.innerHTML;}

function fmt(iso){
  if(!iso) return '';
  const d = new Date(iso);
  if (isNaN(d)) return '';
  return d.toLocaleDateString('en-GB',{day:'numeric',month:'short',year:'numeric'});
}

function visible(){
  return ITEMS.filter(it => {
    if (fCat !== 'all' && !(it.categories||[]).includes(fCat)) return false;
    if (fAuth !== 'all' && it.authority !== fAuth) return false;
    if (fNew && !it.is_new) return false;
    if (fQ) {
      const hay = ((it.title||'') + ' ' + (it.summary||'') + ' ' + (it.source||'')).toLowerCase();
      if (!hay.includes(fQ)) return false;
    }
    return true;
  });
}

function render(){
  const list = visible();
  const box = document.getElementById('list');

  const nNew = ITEMS.filter(i=>i.is_new).length;
  const nOff = ITEMS.filter(i=>i.authority==='official').length;
  document.getElementById('stats').innerHTML =
    `<div class="stat"><div class="n">${ITEMS.length}</div><div class="l">Total tracked</div></div>
     <div class="stat"><div class="n">${nNew}</div><div class="l">New (${NEWDAYS} days)</div></div>
     <div class="stat"><div class="n">${nOff}</div><div class="l">Official sources</div></div>
     <div class="stat"><div class="n">${list.length}</div><div class="l">Showing</div></div>`;

  if(!list.length){
    box.innerHTML = '<div class="empty">No items match these filters.</div>';
    return;
  }
  box.innerHTML = list.map(it => `
    <div class="item ${it.authority} ${it.is_new?'isnew':''}">
      <h3><a href="${esc(it.url)}" target="_blank" rel="noopener">${esc(it.title)}</a></h3>
      ${it.summary ? `<div class="sum">${esc(it.summary.slice(0,260))}${it.summary.length>260?'…':''}</div>` : ''}
      <div class="tags">
        ${it.is_new ? '<span class="badge">NEW</span>' : ''}
        ${it.authority === 'official' ? '<span class="badge off">OFFICIAL</span>' : ''}
        ${(it.categories||[]).map(c=>`<span class="tag">${esc(c)}</span>`).join('')}
        <span class="src">${esc(it.source)} · ${esc(it.domain)}${it.published?' · '+fmt(it.published):' · seen '+fmt(it.first_seen)}</span>
      </div>
    </div>`).join('');
}

document.addEventListener('click', e => {
  const b = e.target.closest('button.f');
  if(!b) return;
  if(b.dataset.cat !== undefined){
    fCat = b.dataset.cat;
    catBox.querySelectorAll('button').forEach(x=>x.classList.toggle('on', x===b));
  } else if(b.dataset.auth !== undefined){
    fAuth = b.dataset.auth;
    b.parentElement.querySelectorAll('button[data-auth]').forEach(x=>x.classList.toggle('on', x===b));
  } else if(b.dataset.new !== undefined){
    fNew = !fNew;
    b.classList.toggle('on', fNew);
  }
  render();
});

document.getElementById('q').addEventListener('input', e => {
  fQ = e.target.value.toLowerCase().trim();
  render();
});

render();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="UAE tax and reporting update monitor")
    ap.add_argument("--check", action="store_true", help="test every source URL")
    ap.add_argument("--rebuild", action="store_true", help="rebuild dashboard only")
    ap.add_argument("--digest", action="store_true", help="print the digest of new items")
    ap.add_argument("--export", action="store_true", help="write an Excel export")
    ap.add_argument("--no-email", action="store_true", help="never send email")
    args = ap.parse_args()

    cfg = load_config()

    if args.check:
        check_sources(cfg)
        return 0

    store = load_store()

    if args.rebuild:
        n = build_dashboard(cfg, store)
        print(f"  Dashboard rebuilt with {n} items")
        return 0

    if args.export:
        export_excel(store)
        return 0

    print(f"\n  UAE Tax Monitor — {datetime.now().strftime('%d %b %Y %H:%M')}")
    print("  " + "-" * 62)
    new_items, report = run(cfg, store)

    for name, status, detail in report:
        mark = "  ok  " if status == "OK" else "  !!  "
        print(f"{mark}{name:<44} {detail}")
    print("  " + "-" * 62)

    save_store(store)
    total = build_dashboard(cfg, store)
    print(f"  {len(new_items)} new item(s). {total} tracked in total.")

    digest = build_digest(new_items, cfg)
    if digest:
        with open(os.path.join(DATA_DIR, "latest_digest.txt"), "w", encoding="utf-8") as fh:
            fh.write(digest)
        if args.digest:
            print("\n" + digest)
        if not args.no_email:
            send_email(digest, len(new_items))

    # expose the count to GitHub Actions
    gh_out = os.environ.get("GITHUB_OUTPUT")
    if gh_out:
        with open(gh_out, "a", encoding="utf-8") as fh:
            fh.write(f"new_count={len(new_items)}\n")

    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
